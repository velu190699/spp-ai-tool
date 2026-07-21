import re
from pathlib import Path

from openpyxl import load_workbook

from src.settlement.settlement_report import build, story_items


def test_extraction_prompt_declares_version_one_story_and_fidelity():
    # The prompt is the reproducibility contract: version tag (stamped into every
    # stories JSON), the ONE STORY PER RR rule, and the fidelity/no-editorial rule.
    text = Path("src/settlement/rr_extraction_prompt.md").read_text(encoding="utf-8")
    assert re.match(r"PROMPT_VERSION:\s*\S+", text)
    assert "ONE STORY PER RR" in text
    assert "FIDELITY, NOT JUDGMENT" in text


def test_rewrite_item_pages_sets_per_determinant_pages():
    from src.settlement.settlement_report import rewrite_item_pages

    desc = (
        "a. Add a parameter for go live of this RR (RR728). Go Live Date TBD.\n\n"
        "1. Update the calculation for #RtMwpDistHrlyAmt to match: X. [p.15]\n"
        "2. Update the calculation for RtDevHrlyQty to match: Y. [p.15]\n"
        "3. Add RtVirtReplaceHrlyQty: Z. [p.15]\n\n"
        "Background: Tariff Attachment AE 8.6.7. [p.15]\n\n"
        "Settlement User Guide: https://example/protocols"
    )
    page_by_det = {"RtMwpDistHrlyAmt": 15, "RtDevHrlyQty": 16, "RtVirtReplaceHrlyQty": 18}
    out = rewrite_item_pages(desc, page_by_det)

    assert "#RtMwpDistHrlyAmt to match: X. [p.15]" in out
    assert "RtDevHrlyQty to match: Y. [p.16]" in out
    assert "RtVirtReplaceHrlyQty: Z. [p.18]" in out
    # Go-live block and trailing Background paragraph keep their original text.
    assert "go live of this RR (RR728)" in out
    assert "Background: Tariff Attachment AE 8.6.7. [p.15]" in out


def test_rewrite_item_pages_leaves_unmapped_items_untouched():
    from src.settlement.settlement_report import rewrite_item_pages

    desc = "1. Update the calculation for #Unknown to match: X. [p.15]\n"
    assert rewrite_item_pages(desc, {"RtDevHrlyQty": 16}) == desc

SAMPLE_REPORT = {
    "rr_id": "RR900",
    "rr_title": "Sample Settlement Test RR",
    "sharepoint_url": "https://sp/RR900",
    "total_pages": 12,
    "rr_class": "SETTLEMENT_CALC",
    "charge_type_index": [
        {"banner": "Market Protocols / Settlement User Guide", "section": "4.5.12",
         "title": "Revenue Neutrality Uplift Distribution Amount", "is_new": False, "page": 3},
        {"banner": "Market Protocols / Settlement User Guide", "section": "4.5.19",
         "title": "SSR Distribution Amount", "is_new": True, "page": 4},
    ],
    "reconciliation": {"status": "PASS"},
}


def test_build_writes_rr_summary_and_settlement_stories_sheets(tmp_path):
    out_path = tmp_path / "SPP_RR_Report_Summary.xlsx"
    results = [{"report": SAMPLE_REPORT, "stories": None}]

    build(results, str(out_path))

    wb = load_workbook(out_path)
    assert wb.sheetnames == ["RR Summary", "Settlement Stories"]

    summary_row = next(wb["RR Summary"].iter_rows(min_row=3, max_row=3, values_only=True))
    assert summary_row[0] == "RR900"
    assert summary_row[2] == "Settlement Calc"
    assert summary_row[3] == "Pass"

    stories_rows = list(wb["Settlement Stories"].iter_rows(min_row=3, values_only=True))
    assert len(stories_rows) == 2  # no LLM stories -> fallback: one row per charge type
    assert all("no LLM story yet" in row[3] for row in stories_rows)
    assert {row[5] for row in stories_rows} == {"MODIFIED", "ADDED"}


SAMPLE_STORIES = {
    "jira_stories": [{
        "summary": "RR900 §4.5.19 – Add SSR Distribution Amount",
        "description": (
            "a. Add a parameter for go live of this RR (RR900). Go Live Date TBD.\n\n"
            "1. Update the calculation for #RevNeutUpliftDistAmt to match: X * Y (was: X). [p.3]\n"
            "2. Add calculation for #SsrMnthlyDistAoAmt: SUM_s ( SsrMnthlyAmt * SsrShareAoPct ) * (-1).\n"
            "Rounded to 2 decimals. [p.4]\n"
            "3. Remove RtOldDevHrlyQty from the deviation calculation — deleted in this RR. [p.4]\n\n"
            "Background: Tariff Attachment AE has the parallel change.\n\n"
            "Settlement User Guide: https://example/protocols"
        ),
    }],
}


def test_build_writes_one_settlement_stories_row_per_story_item(tmp_path):
    out_path = tmp_path / "SPP_RR_Report_Summary.xlsx"
    results = [{"report": SAMPLE_REPORT, "stories": SAMPLE_STORIES}]

    build(results, str(out_path))

    rows = list(load_workbook(out_path)["Settlement Stories"].iter_rows(min_row=3, values_only=True))
    assert len(rows) == 3  # one row per numbered item, not per charge type
    assert [row[2] for row in rows] == [1, 2, 3]
    assert rows[0][4] == "#RevNeutUpliftDistAmt" and rows[0][5] == "MODIFIED"
    assert rows[1][4] == "#SsrMnthlyDistAoAmt" and rows[1][5] == "ADDED"
    assert "Rounded to 2 decimals" in rows[1][3]  # wrapped lines folded into the item
    assert rows[2][4] == "RtOldDevHrlyQty" and rows[2][5] == "DELETED"
    assert rows[0][6] == "RR900 Rec. Report, p.3"
    assert all(row[1] == "RR900 §4.5.19 – Add SSR Distribution Amount" for row in rows)


def test_item_determinant_handles_acronym_runs():
    from src.settlement.settlement_report import _item_determinant

    # "URD" broke the old camel-hump regex, so the column showed the containing
    # formula (RtRevInadqcSppAmt) instead of the determinant the item is about.
    assert _item_determinant(
        "Add RtURDMpAmt as a new additive term in the Real-Time Revenue "
        "Inadequacy Amount (RtRevInadqcSppAmt) formula. [p.53]"
    ) == "RtURDMpAmt"
    assert _item_determinant("Update #RtCalMtr5minQty to match: X. [p.4]") == "#RtCalMtr5minQty"
    assert _item_determinant("No determinant in this sentence.") == ""


def test_story_items_excludes_go_live_block_and_trailing_paragraphs():
    items = story_items(SAMPLE_STORIES["jira_stories"][0]["description"])
    assert [n for n, _ in items] == [1, 2, 3]
    text = " ".join(t for _, t in items)
    assert "go live" not in text and "Background" not in text and "https" not in text


def test_build_skips_settlement_stories_for_non_calc_rrs(tmp_path):
    out_path = tmp_path / "SPP_RR_Report_Summary.xlsx"
    non_calc_report = dict(SAMPLE_REPORT, rr_class="TARIFF_GOVERNANCE", charge_type_index=[],
                            reconciliation={"status": "NO_CHARGE_CODES"})
    results = [{"report": non_calc_report, "stories": None}]

    build(results, str(out_path))

    wb = load_workbook(out_path)
    stories_rows = list(wb["Settlement Stories"].iter_rows(min_row=3, values_only=True))
    assert stories_rows == []


def test_build_escapes_leading_equals_sign_to_prevent_formula_injection(tmp_path):
    out_path = tmp_path / "SPP_RR_Report_Summary.xlsx"
    malicious_report = dict(SAMPLE_REPORT, rr_title="=cmd|'/c calc'!A1")
    results = [{"report": malicious_report, "stories": None}]

    build(results, str(out_path))

    wb = load_workbook(out_path)
    title_cell = list(wb["RR Summary"].iter_rows(min_row=3, max_row=3, values_only=True))[0][1]
    assert not title_cell.startswith("=")
