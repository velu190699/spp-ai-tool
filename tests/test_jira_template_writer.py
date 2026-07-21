from pathlib import Path

from openpyxl import load_workbook

from src.settlement.jira_template_writer import (
    StoryRow,
    stories_from_results,
    write_story_workbook,
)

TEMPLATE = Path("templates/Jira_Story_Creator_template.xlsx")


def _result(rr, cls, status, *, stories=None, index=None, initiative="", citation="",
            version="", url=""):
    return {
        "report": {
            "rr_id": rr,
            "rr_class": cls,
            "reconciliation": {"status": status},
            "charge_type_index": index or [],
            "sharepoint_url": url,
            "market_initiative": initiative,
            "market_initiative_citation": citation,
            "protocol_version": version,
        },
        "stories": stories,
    }


def test_stories_have_no_epic_and_blank_epic_column():
    index = [
        {"banner": "Market Protocols", "section": "4.5.18", "title": "RUC MWP Distribution", "page": 12, "is_new": True},
        {"banner": "Market Protocols", "section": "4.5.12", "title": "Revenue Neutrality Uplift", "page": 9, "is_new": False},
        {"banner": "Tariff", "section": "1.1", "title": "Definitions", "page": 2, "is_new": False},
    ]
    rows = stories_from_results([_result("RR728", "SETTLEMENT_CALC", "PASS", index=index)])

    # No epic row; PM fills the Epic column — every row's epic is blank.
    assert all(r.issue_type != "Epic" for r in rows)
    assert all(r.epic == "" for r in rows)
    assert all(r.create == "" for r in rows)  # nothing pre-approved
    # One story per Market-Protocols charge entry (Tariff entry excluded).
    assert len(rows) == 2
    assert all(r.summary.startswith("[SPPIM: Back Office: Settlements]") for r in rows)


def test_footer_carries_verbatim_initiative_citation_and_version():
    index = [{"banner": "Market Protocols", "section": "4.5.18", "title": "RUC MWP", "page": 12, "is_new": True}]
    rows = stories_from_results([_result(
        "RR728", "SETTLEMENT_CALC", "PASS", index=index,
        initiative="2026 Settlements Fall Bundle",
        citation="(07) Settlement Releases - CUF July 2026.pdf:p6",
        version="112a",
    )])
    desc = rows[0].description
    assert "2026 Settlements Fall Bundle [(07) Settlement Releases - CUF July 2026.pdf:p6]" in desc
    assert "Settlement User Guide) version: 112a" in desc


def test_review_tasks_and_out_of_scope():
    rows = stories_from_results([
        _result("RR720", "SETTLEMENT_CALC", "HARD_FAIL"),
        _result("RR773", "SETTLEMENT_RELEVANT", "REVIEW_SETTLEMENT_PROSE"),
        _result("RR665", "TARIFF_GOVERNANCE", "NO_CHARGE_CODES"),
    ])
    assert sum(1 for r in rows if "RR720" in r.summary and r.issue_type == "Task") == 1
    assert sum(1 for r in rows if "RR773" in r.summary and r.issue_type == "Task") == 1
    assert not any("RR665" in r.summary for r in rows)  # TARIFF_GOVERNANCE emits nothing


def test_stories_prefer_llm_output_when_present():
    llm = {"jira_stories": [
        {"summary": "RR623 §4.5.19 – Add SSR Distribution Amount", "description": "Add the calc.",
         "acceptance_criteria": ["#SsrMnthlyDistAoAmt computed per formula"]},
    ]}
    rows = stories_from_results([_result("RR623", "SETTLEMENT_CALC", "PASS", stories=llm, url="https://sp/rr623")])
    story = rows[0]
    assert story.summary.startswith("[SPPIM: Back Office: Settlements] RR623")
    assert "Add the calc." in story.description
    assert "https://sp/rr623" in story.description  # RR footer appended
    assert story.acceptance_criteria.startswith("# #SsrMnthlyDistAoAmt")
    assert "DST dates" in story.acceptance_criteria  # boilerplate follows


def test_workbook_description_uses_item_codes_instead_of_formulas():
    from src.settlement.jira_template_writer import _workbook_description

    story = {
        "description": (
            "a. Add a parameter for go live of this RR (RR728). Go Live Date TBD.\n\n"
            "1. Update the calculation for #RtMwpDistHrlyAmt to match: A = B * C. [p.16]\n"
            "2. Add calculation for #RtDevHrlyQty: D = E + F. [p.17]\n\n"
            "Background: Tariff Attachment AE.\n\n"
            "Recommendation Report (SharePoint): https://sp/rr728"
        ),
        "items": [
            {"n": 1, "action": "Update the calculation for #RtMwpDistHrlyAmt (BAA level).",
             "determinant": "#RtMwpDistHrlyAmt", "code": "RR728-01", "page": 16},
            {"n": 2, "action": "Add the calculation for #RtDevHrlyQty.",
             "determinant": "#RtDevHrlyQty", "code": "RR728-02", "page": 17},
        ],
    }
    out = _workbook_description(story, "\n\nFOOTER")

    assert "1. Update the calculation for #RtMwpDistHrlyAmt (BAA level). [RR728-01] [p.16]" in out
    assert "2. Add the calculation for #RtDevHrlyQty. [RR728-02] [p.17]" in out
    assert "A = B * C" not in out and "D = E + F" not in out  # formulas replaced by codes
    assert "go live of this RR" in out  # go-live block kept
    assert "Background: Tariff Attachment AE." in out  # trailing paragraph kept
    assert out.endswith("FOOTER")


def test_workbook_description_splits_codes_for_multi_screenshot_items():
    # An item whose formula spans two screenshots lists both codes; one with no
    # crop (parts == 0) shows no code; parts unset behaves like a single code.
    from src.settlement.jira_template_writer import _workbook_description

    story = {
        "description": (
            "1. Update #A. [p.16]\n"
            "2. Update #B. [p.17]\n"
            "3. Update #C. [p.18]\n\n"
            "Background."
        ),
        "items": [
            {"n": 1, "action": "Update #A.", "code": "RR728-01", "page": 16, "parts": 2},
            {"n": 2, "action": "Update #B.", "code": "RR728-02", "page": 17, "parts": 0},
            {"n": 3, "action": "Update #C.", "code": "RR728-03", "page": 18},
        ],
    }
    out = _workbook_description(story, "")

    assert "1. Update #A. [RR728-01a] [RR728-01b] [p.16]" in out
    assert "2. Update #B. [p.17]" in out and "RR728-02" not in out  # no crop -> no code
    assert "3. Update #C. [RR728-03] [p.18]" in out                  # parts unset -> one code


def test_workbook_description_without_items_is_unchanged():
    from src.settlement.jira_template_writer import _workbook_description

    assert _workbook_description({"description": "Plain body."}, " + F") == "Plain body. + F"


def test_no_relevant_rrs_returns_empty():
    assert stories_from_results([_result("RR665", "TARIFF_GOVERNANCE", "NO_CHARGE_CODES")]) == []


def test_write_story_workbook_respects_template_contract(tmp_path):
    rows = [
        StoryRow(summary="[SPPIM: Back Office: Settlements] RR728 — §4.5.18 RUC MWP (Added)",
                 description="Story body.", acceptance_criteria="# AC"),
    ]
    out = tmp_path / "out.xlsx"
    write_story_workbook(TEMPLATE, out, rows)

    wb = load_workbook(out)
    ws = wb["Jira Stories"]
    headers = {str(ws.cell(row=4, column=c).value or "").strip(): c for c in range(1, ws.max_column + 1)}

    # Example rows replaced: exactly header + our row remain.
    assert ws.max_row == 4 + len(rows)
    assert ws.cell(row=5, column=headers["Issue Type"]).value == "Story"
    assert ws.cell(row=5, column=headers["Epic"]).value is None  # PM assigns the epic
    assert ws.cell(row=5, column=headers["Create?"]).value is None  # PM opts in

    # Green columns (the sync app's) untouched.
    for green in ("Jira Key", "Sync Status", "Sync Timestamp", "Sync Error"):
        assert ws.cell(row=5, column=headers[green]).value is None

    # The Tests sheet survives, but its shipped example rows are removed —
    # we don't author tests, so the tab is blank below its header.
    assert "Tests" in wb.sheetnames
    tests_ws = wb["Tests"]
    tests_header = next(
        row for row in range(1, tests_ws.max_row + 1)
        if str(tests_ws.cell(row=row, column=1).value or "").strip() == "Create?"
    )
    assert tests_ws.max_row == tests_header
    # The source template still has its example rows (we filled a copy).
    assert load_workbook(TEMPLATE)["Jira Stories"].max_row > 4
    assert load_workbook(TEMPLATE)["Tests"].max_row > 4


def test_write_story_workbook_adds_screenshot_tab(tmp_path):
    # Per Miquel's guide: Local ID in the row, sheet named EXACTLY like it,
    # screenshots on that sheet. Images land in the xlsx package's media store.
    from PIL import Image as PILImage

    png = tmp_path / "page-015.png"
    PILImage.new("RGB", (1200, 800), "white").save(png)

    row = StoryRow(summary="[SPPIM: Back Office: Settlements] RR728 test", local_id="RR728")
    out = tmp_path / "out.xlsx"
    write_story_workbook(TEMPLATE, out, [row],
                         screenshots={"RR728": [(png, "RR728 Recommendation Report — p.15")]})

    wb = load_workbook(out)
    assert "RR728" in wb.sheetnames
    assert wb["RR728"]["A1"].value == "RR728 Recommendation Report — p.15"

    import zipfile
    with zipfile.ZipFile(out) as z:
        media = [n for n in z.namelist() if n.startswith("xl/media/")]
    assert media, "screenshot image missing from the xlsx package"

    # Local ID written to column C of the story row.
    ws = wb["Jira Stories"]
    headers = {str(ws.cell(row=4, column=c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=5, column=headers["Local ID"]).value == "RR728"
