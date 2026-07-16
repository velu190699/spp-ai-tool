#!/usr/bin/env python3
"""
settlement_report.py — writes SPP_RR_Report_Summary.xlsx from a list of results.

Technical: each result is {"report": <rr_structure.extract() JSON>, "stories":
<Claude JSON or None>}. Produces two sheets:
  Sheet 1 "RR Summary"        — one row per RR: class, status, sections+pages,
                                citation (RR + pages), clickable SharePoint link.
  Sheet 2 "Settlement Stories"— one row per numbered change item in the LLM
                                stories (the "1. Update the calculation for..."
                                items), with the item's determinant, action and
                                page citation. When an RR has no LLM stories yet
                                it falls back to one row per charge type.

Business: this is the settlement-development team's intake queue, not the PCI
market-changes briefing (see src/summaries/). Every row here is scoped to what a
developer needs to change code: which charge type moved, whether it's new or
modified, and a page-anchored citation so a reviewer can verify against the
source RR without re-reading the whole document.
"""
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def _bd():
    s=Side(style="thin",color="D0D0D0"); return Border(left=s,right=s,top=s,bottom=s)
def _W(h="left"): return Alignment(horizontal=h,vertical="top",wrap_text=True)
def _C(): return Alignment(horizontal="center",vertical="top",wrap_text=True)

CLASS_STYLE={
 "SETTLEMENT_CALC":("EAF3DE","27500A","Extract charge-type stories"),
 "SETTLEMENT_RELEVANT":("FAEEDA","633806","Manual review — settlement impact in Tariff prose"),
 "TARIFF_GOVERNANCE":("F1EFE8","444441","Out of scope for calc stories"),
}
STATUS_STYLE={"PASS":("C0DD97","27500A"),"PASS_NUMBERING_MAPPED":("C0DD97","27500A"),
 "REVIEW_SETTLEMENT_PROSE":("FAC775","633806"),
 "NO_CHARGE_CODES":("F1EFE8","444441"),"HARD_FAIL":("F7C1C1","791F1F")}
CT_STATUS={"ADDED":("C0DD97","27500A"),"MODIFIED":("FAC775","633806"),"DELETED":("F7C1C1","791F1F")}

# One "change item" = one numbered entry in a story description ("1. Update the
# calculation for #X to match: ... [p.15]"). The description format is pinned by
# rr_extraction_prompt.md, so the parser here only has to split on the numbers.
_ITEM_BOUNDARY=re.compile(r"(?m)^\s*(\d+)\.\s+")
_PAGE_REF=re.compile(r"\[p\.\s*([^\]]+)\]")
_HASH_CODE=re.compile(r"#\w+")
# Determinants without their '#' (RtDevHrlyQty, SsrShareAoPct, RtURDMpAmt):
# >=3 CamelCase humps, where a hump may start with an acronym run ("URD", "DC").
_BARE_CODE=re.compile(r"\b(?:[A-Z]+[a-z0-9]+){3,}\b")
_ACTION_BY_VERB={"add":"ADDED","update":"MODIFIED","replace":"MODIFIED",
                 "remove":"DELETED","delete":"DELETED"}


def story_items(description):
    """Split a story description into its numbered change items.

    Returns [(item_number, single_line_text), ...]. The go-live block ("a. Add a
    parameter...") and the trailing Background/link paragraphs are not numbered
    items and are excluded; the last item is cut at the first blank line.
    """
    marks=list(_ITEM_BOUNDARY.finditer(description or ""))
    items=[]
    for k,m in enumerate(marks):
        end=marks[k+1].start() if k+1<len(marks) else len(description)
        text=description[m.end():end].split("\n\n")[0]
        items.append((int(m.group(1))," ".join(text.split())))
    return items


def _item_action(text):
    verb=text.split(None,1)[0].lower().rstrip(":") if text else ""
    return _ACTION_BY_VERB.get(verb,"MODIFIED")


def _item_determinant(text):
    match=_HASH_CODE.search(text) or _BARE_CODE.search(text)
    return match.group(0) if match else ""


def build(results, out_path):
    wb=Workbook()
    ws=wb.active; ws.title="RR Summary"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:H1")
    t=ws["A1"]; t.value="SPP RR Analysis — Report Summary (with citations)"
    t.font=Font(name="Arial",size=13,bold=True,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="2C2C2A")
    t.alignment=_W(); ws.row_dimensions[1].height=26
    cols=["RR ID","Title","Class","Status","Charge Types / Sections (page)","Citation","SharePoint Link","Market Initiative"]
    wid=[8,28,18,22,34,28,26,18]
    for ci,c in enumerate(cols,1):
        cell=ws.cell(row=2,column=ci,value=c)
        cell.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="444441")
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=_bd()
        ws.column_dimensions[cell.column_letter].width=wid[ci-1]
    ws.row_dimensions[2].height=28

    settlement=[]
    r=3
    for res in results:
        d=res["report"]; cls=d["rr_class"]; status=d["reconciliation"]["status"]
        rr=d.get("rr_id","?"); title=d.get("rr_title","") or ""
        url=d.get("sharepoint_url") or ""
        # Only Market Protocols (a.k.a. Settlement User Guide) sections matter to
        # the settlement team — Tariff/Planning Criteria/etc. are not listed.
        mp=[i for i in d["charge_type_index"] if i["banner"].startswith("Market")]
        secs="\n".join(f"§{i['section']} {i['title'][:28]} (p.{i.get('page','?')})" for i in mp) if mp else \
             ("— (prose change)" if cls=="SETTLEMENT_RELEVANT" else "— (none)")
        ver=d.get("protocol_version","")
        if ver: secs=f"Settlement User Guide v{ver}\n{secs}"
        pages=sorted({i.get("page") for i in mp if i.get("page")})
        cite=f"{rr} Recommendation Report" + (f", pp. {', '.join(map(str,pages))}" if pages else f", {d.get('total_pages','?')} pp.")
        cbg,cfg,_=CLASS_STYLE.get(cls,("F1EFE8","444441","")); sbg,sfg=STATUS_STYLE.get(status,("F1EFE8","444441"))
        # Initiative is the VERBATIM slide wording; the citation line under it
        # names the file/page so the claim is checkable against the slide.
        initiative=d.get("market_initiative","")
        if initiative and d.get("market_initiative_citation"):
            initiative=f"{initiative}\n({d['market_initiative_citation']})"
        vals=[rr,title,cls.replace("_"," ").title(),status.replace("_"," ").title(),secs,cite,url,initiative]
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=r,column=ci,value=val); c.border=_bd(); c.font=Font(name="Arial",size=10); c.alignment=_W()
            if ci==1: c.font=Font(name="Arial",size=10,bold=True,color="185FA5"); c.alignment=_C()
            if ci==3: c.fill=PatternFill("solid",fgColor=cbg); c.font=Font(name="Arial",size=10,bold=True,color=cfg); c.alignment=_C()
            if ci==4: c.fill=PatternFill("solid",fgColor=sbg); c.font=Font(name="Arial",size=10,bold=True,color=sfg); c.alignment=_C()
            if ci==5: c.font=Font(name="Courier New",size=10)
            if ci==7 and url:
                c.value="Open in SharePoint"; c.hyperlink=url
                c.font=Font(name="Arial",size=10,color="185FA5",underline="single")
        ws.row_dimensions[r].height=72; r+=1
        if cls=="SETTLEMENT_CALC":
            settlement.append((rr,url,mp,res.get("stories")))

    ws.freeze_panes="A3"

    ws2=wb.create_sheet("Settlement Stories"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:H1")
    t=ws2["A1"]; t.value="Settlement Stories — one row per change item (numbered items from the stories; one row per charge type when no story yet)"
    t.font=Font(name="Arial",size=12,bold=True,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="2C2C2A")
    t.alignment=_W(); ws2.row_dimensions[1].height=24
    c2=["RR ID","Story Summary","Item #","Change Item","Determinant","Action","Citation (RR + page)","SharePoint Link"]
    w2=[8,36,7,64,20,12,20,18]
    for ci,c in enumerate(c2,1):
        cell=ws2.cell(row=2,column=ci,value=c)
        cell.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="444441")
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=_bd()
        ws2.column_dimensions[cell.column_letter].width=w2[ci-1]
    ws2.row_dimensions[2].height=26

    def _emit(r,vals,url,action,tall):
        for ci,val in enumerate(vals,1):
            c=ws2.cell(row=r,column=ci,value=val); c.border=_bd(); c.font=Font(name="Arial",size=10); c.alignment=_W()
            if ci==1: c.font=Font(name="Arial",size=10,bold=True,color="185FA5"); c.alignment=_C()
            if ci==3: c.alignment=_C()
            if ci==4: c.font=Font(name="Courier New",size=10)
            if ci==5: c.font=Font(name="Arial",size=10,bold=True)
            if ci==6 and action in CT_STATUS:
                bg,fg=CT_STATUS[action]; c.fill=PatternFill("solid",fgColor=bg)
                c.font=Font(name="Arial",size=10,bold=True,color=fg); c.alignment=_C()
            if ci==8 and url:
                c.value="Open"; c.hyperlink=url; c.font=Font(name="Arial",size=10,color="185FA5",underline="single"); c.alignment=_C()
        ws2.row_dimensions[r].height=64 if tall else 32

    r=3
    for rr,url,mp,stories in settlement:
        llm_stories=(stories or {}).get("jira_stories") if isinstance(stories,dict) else None
        if llm_stories:
            for s in llm_stories:
                summary=s.get("summary","")
                items=story_items(s.get("description",""))
                for num,text in items:
                    page=_PAGE_REF.search(text)
                    cite=f"{rr} Rec. Report, p.{page.group(1)}" if page else f"{rr} Rec. Report"
                    action=_item_action(text)
                    _emit(r,[rr,summary,num,text,_item_determinant(text),action,cite,url],url,action,tall=True)
                    r+=1
                if not items:  # story without numbered items — keep it visible anyway
                    _emit(r,[rr,summary,"—"," ".join(str(s.get("description","")).split())[:300],"","",f"{rr} Rec. Report",url],url,"",tall=True)
                    r+=1
        else:
            for i in mp:
                st="ADDED" if i["is_new"] else "MODIFIED"
                summary=f"{rr} – Market Protocols §{i['section']} {i['title']}"
                cite=f"{rr} Rec. Report, p.{i.get('page','?')}"
                _emit(r,[rr,summary,"",f"§{i['section']} {i['title']} — no LLM story yet (run --call-claude)","",st,cite,url],url,st,tall=False)
                r+=1
    ws2.freeze_panes="A3"

    wb.save(out_path)
    # escape leading '='
    wb2=load_workbook(out_path)
    for wsx in wb2.worksheets:
        for row in wsx.iter_rows():
            for c in row:
                if isinstance(c.value,str) and c.value.startswith("="): c.value=" "+c.value
    wb2.save(out_path)
    return out_path
