"""jira_template_writer.py — fill Miquel's Jira_Story_Creator_template.xlsx.

Technical: loads a COPY of the team's story-creator template (v1.1.0) and
writes Epic/Story rows into the "Jira Stories" sheet. The template's contract:
header row starts with "Create?"; amber columns (Create? .. Acceptance
Criteria) are authored here; green columns (Jira Key, Sync Status, Sync
Timestamp, Sync Error) belong to Miquel's sync app and are NEVER written. The
shipped example rows below the header are replaced with ours.

Business: this is the PM-review artifact — the tool proposes stories, the PM
edits and flips "Create?" to Y, and only then does the sync app create Jira
issues. Every row therefore defaults Create? to blank: the tool must not be
able to push a story into Jira on its own. The epic's Parent Link (initiative
key, e.g. PM-944) is also left blank for the PM.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

_ITEM_BOUNDARY = re.compile(r"(?m)^\s*\d+\.\s+")

from openpyxl import load_workbook

LOGGER = logging.getLogger(__name__)

STORIES_SHEET = "Jira Stories"
TESTS_SHEET = "Tests"
SUMMARY_PREFIX = "[SPPIM: Back Office: Settlements]"
_SCREENSHOT_MAX_WIDTH_PX = 1000

# Header text -> StoryRow attribute. Green columns are deliberately absent.
_FIELD_BY_HEADER = {
    "Create?": "create",
    "Issue Type": "issue_type",
    "Local ID": "local_id",
    "Summary": "summary",
    "Description": "description",
    "Steps to Reproduce": "steps_to_reproduce",
    "Client Ticket": "client_ticket",
    "Epic": "epic",
    "Epic Name": "epic_name",
    "Parent Link": "parent_link",
    "Complete After": "complete_after",
    "Priority": "priority",
    "Acceptance Criteria": "acceptance_criteria",
}

# The team's standing acceptance-criteria items (as seen in SP-12813/SP-12814);
# story-specific criteria from the LLM are prepended to these.
BOILERPLATE_AC = (
    "# DST dates have been impact tested (if applicable)\n"
    "# Changes of scope, decision points, newly identified details, etc. have been documented in the story.\n"
    "# Release Documentation is detailed and accurate. For large and/or significant changes, "
    "ensure Product Manager and/or SME has reviewed."
)


@dataclass
class StoryRow:
    issue_type: str = "Story"
    summary: str = ""
    description: str = ""
    acceptance_criteria: str = ""
    local_id: str = ""
    epic: str = ""
    epic_name: str = ""
    parent_link: str = ""
    priority: str = "Medium"
    steps_to_reproduce: str = ""
    client_ticket: str = ""
    complete_after: str = ""
    create: str = ""  # blank: the PM flips to Y after review — never pre-filled


def _find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 30) + 1):
        if str(ws.cell(row=row, column=1).value or "").strip() == "Create?":
            return row
    raise ValueError(
        f"Sheet '{ws.title}' has no 'Create?' header row — template layout changed; "
        "update jira_template_writer.py against the new template version"
    )


def write_story_workbook(
    template_path: Path,
    out_path: Path,
    rows: list[StoryRow],
    screenshots: dict[str, list[tuple[Path, str]]] | None = None,
) -> Path:
    """Fill a copy of the template with `rows` and save it to `out_path`.

    The template file itself is never modified. Formatting, data validation,
    the Tests sheet's banner/header, and the green sync columns are preserved
    untouched so Miquel's app sees exactly the workbook shape it expects. The
    Tests sheet's shipped EXAMPLE rows are removed — we don't author tests, so
    the tab ships blank below its header (Elizabeth/Eduardo, 2026-07-16).

    `screenshots` maps a row's Local ID to [(png_path, caption), ...]. Per
    Miquel's story-creation guide, each Local ID gets a sheet with that EXACT
    name holding the images — his sync app attaches them to the Jira issue.
    """
    wb = load_workbook(template_path)
    if STORIES_SHEET not in wb.sheetnames:
        raise ValueError(f"Template has no '{STORIES_SHEET}' sheet (found: {wb.sheetnames})")
    if TESTS_SHEET in wb.sheetnames:
        tests = wb[TESTS_SHEET]
        tests_header = _find_header_row(tests)
        if tests.max_row > tests_header:
            tests.delete_rows(tests_header + 1, tests.max_row - tests_header)
    ws = wb[STORIES_SHEET]
    header_row = _find_header_row(ws)
    columns = {
        str(ws.cell(row=header_row, column=col).value or "").strip(): col
        for col in range(1, ws.max_column + 1)
    }
    missing = [h for h in _FIELD_BY_HEADER if h not in columns]
    if missing:
        raise ValueError(f"Template is missing expected columns: {missing} — layout changed?")

    # Replace the shipped EXAMPLE rows ("Rows 5+ are EXAMPLES") with ours.
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

    for offset, story in enumerate(rows, start=1):
        row = header_row + offset
        for header, attr in _FIELD_BY_HEADER.items():
            value = getattr(story, attr)
            ws.cell(row=row, column=columns[header], value=value or None)

    for local_id, images in (screenshots or {}).items():
        if images:
            _add_screenshot_sheet(wb, local_id, images)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    LOGGER.info("Jira story workbook written: %s (%d rows)", out_path, len(rows))
    return out_path


def _add_screenshot_sheet(wb, local_id: str, images: list[tuple[Path, str]]) -> None:
    """One sheet named EXACTLY local_id, images stacked with caption rows."""
    from openpyxl.drawing.image import Image as XlsxImage

    ws = wb.create_sheet(str(local_id))
    ws.sheet_view.showGridLines = False
    anchor_row = 1
    for png_path, caption in images:
        ws.cell(row=anchor_row, column=1, value=caption)
        image = XlsxImage(str(png_path))
        # Scale to a readable on-screen width; Excel keeps the aspect ratio we set.
        if image.width and image.width > _SCREENSHOT_MAX_WIDTH_PX:
            ratio = _SCREENSHOT_MAX_WIDTH_PX / image.width
            image.width = int(image.width * ratio)
            image.height = int(image.height * ratio)
        ws.add_image(image, f"A{anchor_row + 1}")
        # ~20 px per default row: leave room for the image plus a spacer row.
        anchor_row += int(image.height / 20) + 3


def _rr_footer(rr: str, url: str, initiative: str, initiative_citation: str = "",
               protocol_version: str = "", include_rr_link: bool = True) -> str:
    lines = ["", ""]
    if include_rr_link:
        # Only when the description doesn't already carry the RR link (the
        # pipeline appends it) — avoids a duplicate link in the workbook.
        lines.append(f"Recommendation Report: {rr} Recommendation Report.docx")
        if url:
            lines.append(f"SharePoint: {url}")
    if protocol_version:
        lines.append(f"Market Protocols (Settlement User Guide) version: {protocol_version}")
    if initiative:
        # Verbatim slide wording + its file/page so the claim is checkable.
        cite = f" [{initiative_citation}]" if initiative_citation else ""
        lines.append(f"Market initiative (as named on the slide): {initiative}{cite}")
    return "\n".join(lines)


def _workbook_description(story: dict, footer: str) -> str:
    """Build the Jira workbook description for one story.

    When the story carries the formula-free "items" mirror, each numbered line
    becomes "<n>. <action> [<code>] [p.<page>]" — the screenshot code replaces
    the written formula so Miquel's app inserts the image there (Eduardo,
    2026-07-17). An item whose formula spans two screenshots lists both codes
    ("[…-02a] […-02b]"), so a reviewer sees it has two images; an item whose
    formula could not be cropped (`parts` == 0, e.g. an image-only formula) shows
    no code. The go-live block (before the list) and the trailing Background/link
    paragraphs are preserved from the full description. Without items, the full
    description (formulas and all) is used unchanged.
    """
    description = str(story.get("description", ""))
    items = story.get("items") or []
    marks = list(_ITEM_BOUNDARY.finditer(description))
    if not items or not marks:
        return description + footer
    head = description[:marks[0].start()].rstrip()
    _body, _sep, tail = description[marks[-1].start():].partition("\n\n")
    lines = []
    for item in sorted(items, key=lambda x: int(x.get("n", 0))):
        code = item.get("code") or ""
        page = item.get("page")
        pg = f" [p.{page}]" if page else ""
        lines.append(f"{int(item.get('n', 0))}. {str(item.get('action', '')).strip()}"
                     f"{_code_tags(code, item.get('parts'))}{pg}")
    parts = [p for p in (head, "\n".join(lines), tail.strip()) if p]
    return "\n\n".join(parts) + footer


def _code_tags(code: str, parts: Any) -> str:
    """Screenshot code tag(s) for a description item.

    `parts` is the image count stamped by screenshots.item_screenshots: None
    (screenshots not generated) → one "[code]"; 0 (no crop) → nothing; N >= 2 →
    one tag per image ("[code a]", "[code b]", …), matching the sheet captions.
    """
    if not code:
        return ""
    if parts is None:
        return f" [{code}]"
    n = int(parts)
    if n <= 0:
        return ""
    if n == 1:
        return f" [{code}]"
    return "".join(f" [{code}{chr(ord('a') + i) if i < 26 else str(i + 1)}]" for i in range(n))


def _normalize_ac_item(item: Any) -> str:
    # Strip only a leading "# " list marker — a bare "#" may be a charge-code
    # prefix (e.g. "#SsrMnthlyDistAoAmt") and must survive.
    text = str(item).strip()
    return text[2:].strip() if text.startswith("# ") else text


def _format_ac(items: Any) -> str:
    specific = ""
    if isinstance(items, list) and items:
        specific = "\n".join(f"# {_normalize_ac_item(item)}" for item in items) + "\n"
    return specific + BOILERPLATE_AC


def _prefixed(summary: str) -> str:
    return summary if summary.startswith("[") else f"{SUMMARY_PREFIX} {summary}"


def stories_from_results(results: list[dict[str, Any]]) -> list[StoryRow]:
    """Turn settlement-pipeline results into template rows.

    NO epic row is generated and the Epic column is left blank on every story:
    per-RR workbooks with embedded epics would create duplicate epics in Jira,
    so the PM fills the Epic column with the real epic key (or local ID) during
    review. Granularity within an RR: one Story per charge entry / LLM story.
    LLM-generated stories (SETTLEMENT_CALC + PASS with --call-claude) carry
    their own description/AC; without the LLM a deterministic draft row is
    emitted so the PM still sees the change, marked as needing the full pass.
    SETTLEMENT_CALC + HARD_FAIL and SETTLEMENT_RELEVANT become single review
    Tasks; TARIFF_GOVERNANCE is out of scope and emits nothing.
    """
    story_rows: list[StoryRow] = []
    for res in results:
        d = res["report"]
        cls = d["rr_class"]
        status = d["reconciliation"]["status"]
        rr = d.get("rr_id", "?")
        url = d.get("sharepoint_url") or ""

        def _footer_for(description_text: str) -> str:
            return _rr_footer(
                rr, url,
                d.get("market_initiative", ""),
                d.get("market_initiative_citation", ""),
                d.get("protocol_version", ""),
                include_rr_link=not (url and url in description_text),
            )
        llm = res.get("stories") if isinstance(res.get("stories"), dict) else None
        llm_stories = (llm or {}).get("jira_stories") or []

        if cls == "SETTLEMENT_CALC" and status in ("PASS", "PASS_NUMBERING_MAPPED"):
            if llm_stories:
                for s in llm_stories:
                    story_rows.append(StoryRow(
                        summary=_prefixed(str(s.get("summary", ""))),
                        description=_workbook_description(s, _footer_for(str(s.get("description", "")))),
                        acceptance_criteria=_format_ac(s.get("acceptance_criteria")),
                    ))
            else:
                for entry in (i for i in d["charge_type_index"] if i["banner"].startswith("Market")):
                    st = "ADDED" if entry["is_new"] else "MODIFIED"
                    story_rows.append(StoryRow(
                        summary=_prefixed(f"{rr} — §{entry['section']} {entry['title']} ({st.title()})"),
                        description=(
                            f"As a user, I want the SPPIM settlements calculation updated per {rr} "
                            f"Market Protocols §{entry['section']} {entry['title']} ({st}).\n\n"
                            f"Source: {rr} Recommendation Report, p.{entry.get('page', '?')}."
                            f"{_footer_for('')}\n\n"
                            "DRAFT: generated without LLM story extraction — run "
                            "`settlement-report --call-claude` for full descriptions before PM review."
                        ),
                        acceptance_criteria=_format_ac(None),
                    ))
        elif cls == "SETTLEMENT_CALC":  # HARD_FAIL: extraction couldn't be reconciled
            story_rows.append(StoryRow(
                issue_type="Task",
                summary=_prefixed(f"{rr} — MANUAL REVIEW: charge-code extraction failed reconciliation"),
                description=(
                    f"The automated extraction for {rr} failed its reconciliation gate "
                    f"(status: {status}). A settlements SME must review the RR directly; "
                    "do not trust auto-generated charge-code rows for this RR."
                    f"{_footer_for('')}"
                ),
                acceptance_criteria=_format_ac(["RR reviewed by settlements SME and stories authored manually"]),
            ))
        elif cls == "SETTLEMENT_RELEVANT":
            story_rows.append(StoryRow(
                issue_type="Task",
                summary=_prefixed(f"{rr} — review settlement impact (prose change, no charge-code redlines)"),
                description=(
                    f"{rr} changes settlement-relevant Tariff/Protocol prose without charge-code "
                    "redlines. Review whether PCI settlement logic or documentation is affected."
                    f"{_footer_for('')}"
                ),
                acceptance_criteria=_format_ac(["Impact assessed and documented; follow-up stories created if needed"]),
            ))
        # TARIFF_GOVERNANCE: informational only — no story (Kashmita's scope rule).

    return story_rows
