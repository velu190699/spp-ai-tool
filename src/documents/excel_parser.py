from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.documents.rr_extractor import normalize_rr_number


@dataclass(frozen=True)
class RRRecord:
    rr_number: str
    title: str
    status: str
    link: str
    search_url: str
    primary_working_group: str
    impacted_documents: str
    row: dict[str, Any]


def _clean(value: object) -> str:
    return str(value or "").strip()


def read_open_rrs(path: Path) -> dict[str, RRRecord]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [_clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    open_rrs: dict[str, RRRecord] = {}
    for row_cells in sheet.iter_rows(min_row=2):
        row = {headers[index]: cell.value for index, cell in enumerate(row_cells) if index < len(headers)}
        status = _clean(row.get("Status"))
        if status.lower() != "open":
            continue
        rr_number = normalize_rr_number(row.get("Number"))
        if not rr_number:
            continue
        link_cell = row_cells[1] if len(row_cells) > 1 else None
        search_url = link_cell.hyperlink.target if link_cell is not None and link_cell.hyperlink else ""
        open_rrs[rr_number] = RRRecord(
            rr_number=rr_number,
            title=_clean(row.get("Title")),
            status=status,
            link=_clean(row.get("Link")),
            search_url=search_url,
            primary_working_group=_clean(row.get("Primary Working Group")),
            impacted_documents=_clean(row.get("Impacted Documents")),
            row=row,
        )
    return open_rrs
